import sys
import os
import os.path
import time
import subprocess
import openpyxl
from openpyxl.styles import PatternFill, Font
import requests
import configparser
import re

from PySide6 import QtWidgets
from PySide6.QtWidgets import QMessageBox
from PySide6.QtCore import QProcess
# from PySide6.QtCore import Signal, QObject
from time import sleep
from datetime import datetime
from threading import Timer
from PySide6.QtCore import QTimer

from ui_login import Ui_FormLogin
from ui_programV22 import Ui_FormProgram
from processfiles import *

# url = "http://10.99.250.108:5000/rpc/check_unit_nr"
# data = {"unit_nr":"FCDBMZ","material_nr":"4135195","order_nr":"","hierarchy":"","testcode":""}
# response = requests.post(url, json=data) 
# print("Status Code", response.status_code)
# print("JSON Response ", response.json())

# ----------Create projects folder------------
if not os.path.isdir("projects"):
    os.mkdir("projects")
dir_projects = os.getcwd() + "/projects"
# ----------Create logviews folder------------
if not os.path.isdir("logviews"):
    os.mkdir("logviews")
dir_logviews = os.getcwd() + "/logviews"
# ----------Create firmwares folder------------
if not os.path.isdir("firmwares"):
    os.mkdir("firmwares")
dir_firmwares = os.getcwd() + "/firmwares"
# ----------Create OnlineQDMdata folder------------
if not os.path.isdir("QDMdata"):
    os.mkdir("QDMdata")
OfflineQDMdata = os.getcwd() + "/QDMdata"
# --------------------------------------------
try:
    from ctypes import windll  # Only exists on Windows.
    myappid = 'mycompany.myproduct.subproduct.version'
    windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
except ImportError:
    pass
# A regular expression, to extract the % complete.
progress_re = re.compile("(\d+)%")

def simple_percent_parser(output):
    """
    Matches lines using the progress_re regex,
    returning a single integer for the % progress.
    """
    m = progress_re.search(output)
    if m:
        pc_complete = m.group(1)
        return int(pc_complete)
# --------------------------------------------
class LoginWindow(QtWidgets.QMainWindow, Ui_FormLogin):
    # signal = Signal(str, str, str)
    def __init__(self):
        super(LoginWindow, self).__init__()
        self.setupUi(self)
        # Get the current working directory
        # directory = os.getcwd()
        
        self.fill_listdevice(dir_projects)
        self.pushButton_OK.clicked.connect(self.button_OK)
        self.pushButton_OK.setAutoDefault(True)
        self.pushButton_Cancel.clicked.connect(self.button_Cancel)

        self.lineEdit_LotNo.returnPressed.connect(self.button_OK)

    def button_OK(self):
        MSNV = self.lineEdit_MSNV.text()
        LotNo = self.lineEdit_LotNo.text()
        Project = self.comboBox_Project.currentText()
        QRCode = self.QRCodecheck.isChecked()
        if MSNV != "" and LotNo != "" and Project != "":
            self.window = MainWindow()
            self.window.show()
            self.window.receive_data(MSNV, LotNo, Project, QRCode)
            self.destroy()      
        else:
            self.showDialog("Please input data the first") 

    def button_Cancel(self):
        self.close()

    def fill_listdevice(self, directory):
        files = [f for f in os.listdir(directory) if f.endswith('.ini')]
        for file in files:
            self.comboBox_Project.addItem(file.rsplit('.', 1)[0])
        return self.comboBox_Project.currentText() 

    def showDialog(self, information_data):
        msgBox = QMessageBox()
        msgBox.setIcon(QMessageBox.Information)
        msgBox.setText(information_data)
        msgBox.setWindowTitle("Information / Thông tin")
        msgBox.setStandardButtons(QMessageBox.Ok)
        msgBox.exec()

class MainWindow(QtWidgets.QMainWindow, Ui_FormProgram):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.setupUi(self)
        self.p = None

        self.timerQDM = QTimer()
        self.timerQDM.setInterval(500)
        self.timerQDM.timeout.connect(slot=lambda: self.processTimerQDM())
        self.timerQDM.start()
        
        self.pushButton_Program.clicked.connect(self.button_program)
        self.pushButton_Program.setAutoDefault(True)
        self.lineEdit_QRCode.returnPressed.connect(self.button_program)

    def processTimerQDM(self):
        if os.path.isdir(self.OnlineQDMdata):
            if is_folder_not_empty(OfflineQDMdata):
                copy_files(OfflineQDMdata, self.OnlineQDMdata)
                delete_files(OfflineQDMdata)

    def getFirmwareVersion(self):
        startindex1 = self.FirmwareFile.rfind("\\")
        startindex2 = self.FirmwareFile.rfind("/")
        startindex = max(startindex1, startindex2) + 1
        stopindex = self.FirmwareFile.rfind(".hex")
        if startindex == -1 or stopindex == -1:
            self.FirmwareVersion = "NONE"
            self.pushButton_Program.setEnabled(False)
        else:
            self.FirmwareVersion = self.FirmwareFile[startindex:stopindex]

    def read_Configfile(self):
        # Reading Data
        config = configparser.ConfigParser()
        config.read(self.projectfilepath)

        self.QDMServerURL = config.get("SETTINGS", "QDMServerURL")
        try:    
            self.OnlineQDMdata = config.get("SETTINGS", "OnlineQDMdata")
        except:
            self.OnlineQDMdata = ""
        try:    
            self.NameOfStation = config.get("SETTINGS", "NameOfStation")
        except:
            self.NameOfStation = ""
        try:    
            self.NameOfStep = config.get("SETTINGS", "NameOfStep")
        except:
            self.NameOfStep = ""
        try:    
            self.NameOfVersion = config.get("SETTINGS", "NameOfVersion")
        except:
            self.NameOfVersion = ""

        self.DirProgrammer = config.get("SETTINGS", "DirProgrammer")
        self.ControllerType = config.get("SETTINGS", "ControllerType")
        
        self.CommanderConnect = config.get("SETTINGS", "CommanderConnect")
        self.CommanderProgram = config.get("SETTINGS", "CommanderProgram")
        self.FirmwareFile = config.get("SETTINGS", "FirmwareFile")
        self.getFirmwareVersion()
        try:    
            self.StorageFile = config.get("SETTINGS", "StorageFile")
        except:
            self.StorageFile = ""
        try:    
            self.ExternalLoader = config.get("SETTINGS", "ExternalLoader")
        except:
            self.ExternalLoader = ""
        try:
            self.TestProgram = config.get("SETTINGS", "TestProgram")
        except:
            self.TestProgram = "False"
        try:
            self.QRCodeSave = config.get("SETTINGS", "QRCodeSave")
        except:
            self.QRCodeSave = "True"

        self.timer = QTimer()
        self.timer.timeout.connect(self.recurring_timer)

    def receive_data(self, MSNV, LotNo, Project, QRC_en):
        self.FirstProgram = True
        self.DeviceProgram = Project
        self.label_Project.setText(self.DeviceProgram)
        self.projectfilepath = os.path.join(dir_projects, self.DeviceProgram + ".ini")
        if os.path.isfile(self.projectfilepath):
            self.read_Configfile()
        else:
            self.showDialog("Can't found file "'"Settings_Project.ini"'" in the folder Projects")

        self.Employee = MSNV
        self.LotNumber = LotNo

        if self.QRCodeSave == "True":
            self.QRCenable = True
        else:   
            self.QRCenable = QRC_en

        if self.QRCenable == False:
            self.lineEdit_QRCode.setEnabled(False)
        else:
            # Join the current working directory with the file name
            self.logviewspath = os.path.join(dir_logviews, self.LotNumber + ".xlsx")
            self.detect_create_excel_file()
            if self.TestProgram == "True":
                self.lineEdit_QRCode.setEnabled(False)
    
    def detect_create_excel_file(self):
        if os.path.isfile(self.logviewspath):
            return False
        else:
            workbook = openpyxl.Workbook()

            sheet = workbook.active
            sheet.title = self.LotNumber
            sheet['A1'] = 'QRCode'
            sheet['B1'] = 'Employee'
            sheet['C1'] = 'Result'
            sheet['D1'] = 'Time Programming'
            sheet['E1'] = 'TimeDay'
            sheet['F1'] = 'FirmwareVersion'
            sheet['A1'].font = Font(bold = True)
            sheet['B1'].font = Font(bold = True)
            sheet['C1'].font = Font(bold = True)
            sheet['D1'].font = Font(bold = True)
            sheet['E1'].font = Font(bold = True)
            sheet['F1'].font = Font(bold = True)
            sheet['A1'].fill = PatternFill("solid", start_color="FFA500")
            sheet['B1'].fill = PatternFill("solid", start_color="FFA500")
            sheet['C1'].fill = PatternFill("solid", start_color="FFA500")
            sheet['D1'].fill = PatternFill("solid", start_color="FFA500")
            sheet['E1'].fill = PatternFill("solid", start_color="FFA500")
            sheet['F1'].fill = PatternFill("solid", start_color="FFA500")

            workbook.save(self.logviewspath)
            workbook.close()
            return True
        
    def button_program(self):
        # self.lineEdit_QRCode.setText("testa")
        if self.QRCenable == True and self.TestProgram == "True":
            self.lineEdit_QRCode.setText("Testa")

        self.QRcode = self.lineEdit_QRCode.text()
        self.pushButton_Result.setText("RESULT")
        self.pushButton_Result.setStyleSheet("background-color: none; color: none;")
        self.progressBar.setValue(0)
        self.plainTextResult.clear()
        self.processingCMD = False
        self.resultstout = False
        self.modeProgram = self.QRcode.lower()
        self.Material = self.DeviceProgram[0:7]
        if self.QRCenable == True:
            if self.QRcode == "":
                self.showDialog("Please scan QRCode the first")
            else:
                if self.is_file_open():
                    self.showDialog("File log is open. Please close it first")
                else:
                    # self.modeProgram = self.QRcode.lower()
                    # --------------------------------------------------------------------   
                    if self.modeProgram == "testa" or self.modeProgram == "erase":
                        self.processingCMD = True
                    else:
                        # self.Material = self.DeviceProgram[0:7]
                        self.data = {"unit_nr":self.QRcode,"material_nr":self.Material,"order_nr":"","hierarchy":"","testcode":""}
                        response = requests.post(self.QDMServerURL, json=self.data)

                        # print("Status Code", response.status_code)
                        # print("JSON Response ", response.json())
                        # JSON_SN = response.json()[0]['result']
                        if response.status_code == 200:
                            if response.json()[0]['result'] == "sn_ok":
                                self.processingCMD = True
                            else:
                                self.showDialog("Wrong_material, not exists in QDM")  
        else:
            self.processingCMD = True
        # --------------------------------------------------------------------------------
        if self.processingCMD == True:
            self.pushButton_Result.setText("Processing.....")
            self.timecounter = 0
            if self.FirstProgram == False:
                self.timer.start()
            self.starttime = time.time()

            result = subprocess.run([self.DirProgrammer, self.CommanderConnect], shell=True, capture_output=True, text=True)
            if self.ControllerType in result.stdout:
                self.start_programController()
                self.pushButton_Program.setEnabled(False)   
            else:
                self.timer.stop()
                self.lineEdit_QRCode.clear()
                self.pushButton_Result.setText("RESULT")
                if "No debug probe detected" in result.stdout:
                    self.showDialog("Please check connection Tool Program or Serialnumber")
                elif "No STM32 target found" in result.stdout: 
                    self.showDialog("Please check connection MCU or Power")    
                else:
                    self.showDialog("Please check name of MCU in "'"Settings_Project.ini"'"")

            # if result.returncode == 0:
            #     if self.ControllerType in result.stdout:
            #         self.start_programController()
            #         self.pushButton_Program.setEnabled(False)
            #     else:
            #         self.timer.stop()
            #         self.showDialog("Please check name of MCU in "'"Settings_Project.ini"'"")
            #         self.lineEdit_QRCode.clear()
            #         self.pushButton_Result.setText("RESULT")
            # else:  
            #     self.timer.stop()
            #     self.showDialog("Please check connection Tool Program (Serialnumber) or MCU")
            #     self.lineEdit_QRCode.clear()
            #     self.pushButton_Result.setText("RESULT")
        else:
            self.lineEdit_QRCode.clear()

    def recurring_timer(self):
        if self.timecounter < 98:
            self.timecounter +=1
            self.progressBar.setValue(self.timecounter)

    def start_programController(self):
        if self.p is None:  # No process running.
            self.p = QProcess()  # Keep a reference to the QProcess (e.g. on self) while it's running.
            self.p.readyReadStandardOutput.connect(self.handle_p_stdout)
            self.p.finished.connect(self.process_p_finished)  # Clean up once complete.
            if self.modeProgram == "erase":
                self.p.start(self.DirProgrammer, [self.CommanderConnect, self.CommanderProgram])
            else:
                self.p.start(self.DirProgrammer, [self.CommanderConnect, self.CommanderProgram, self.StorageFile, self.ExternalLoader, self.FirmwareFile])

    def handle_p_stdout(self):
        data = self.p.readAllStandardOutput()
        stdout = bytes(data).decode("utf8", errors='ignore')
        self.plainTextResult.setPlainText(stdout)
        if self.FirstProgram == True:
            # Extract progress if it is in the data.
            progress = simple_percent_parser(stdout)
            if progress:
                self.progressBar.setValue(progress)
        # --------------------------------------------------------------------------------
        if self.modeProgram == "erase":
            if "Mass erase successfully achieved" in stdout:
                self.resultstout = True
            else:
                self.resultstout = False
        else:
            if "File download complete" in stdout or "Download verified successfully" in stdout:
                self.resultstout = True
            else:
                self.resultstout = False

    def process_p_finished(self):
        self.p = None
        # --------------------------------------------------------------------------------
        self.timecounter = 100
        self.sumtime = time.time() - self.starttime
        if self.resultstout == True:
            self.pushButton_Result.setText("PASSED")
            self.pushButton_Result.setStyleSheet("background-color: #4CAF50; color: #fff;")
            self.progressBar.setValue(self.timecounter) 
            # --------------------------------------------------------------------------------
            if self.modeProgram != "erase":
                self.timer.setInterval(int(self.sumtime)*10)
                self.FirstProgram = False
        else:
            self.pushButton_Result.setText("FAILED")
            self.pushButton_Result.setStyleSheet("background-color: #D32F2F; color: #fff;")
            self.progressBar.setValue(0)
        # --------------------------------------------------------------------------------
        if self.QRCenable == True and self.modeProgram != "erase":
            self.Result = self.pushButton_Result.text()
            self.update_excel_file()
            if self.modeProgram != "testa":
                self.updateQDMSystem()
        self.lineEdit_QRCode.clear()
        self.pushButton_Program.setEnabled(True)

    def updateQDMSystem(self):
        # getting the current date and time
        currentdatetime = datetime.now()
        self.timedatum = currentdatetime.strftime("%d/%m/%Y %H:%M:%S")
        self.time_date = currentdatetime.strftime("%d%m%Y%H%M%S")

        self.qdmsender       = "sender         : " + self.NameOfStation
        self.qdmdatum        = "datum          : " + self.timedatum
        self.qdmlaenge       = "laenge         : " 
        self.qdmprotokoll    = "protokoll      : LEIKA"
        self.minus           = "----------------------------------------------------"
        self.KOPF            = "#KOPF {"
        self.BARCODE         = "  #BARCODE = \"" + self.QRcode + "\";"
        self.MATERIAL        = "  #MATERIAL = \"" + self.Material + "\";"
        self.PRUEFCODE       = "  #PRUEFCODE = \"" + self.NameOfStep + "\";"
        self.PRUEFMETHODE    = "  #PRUEFMETHODE = \"N\";"
        self.PRUEFFLOWCODE   = "  #PRUEFFLOWCODE = \"" + self.NameOfVersion + "\";"
        self.PLATZNUMMER     = "  #PLATZNUMMER = \"" + self.NameOfStation[3:6] + "\";"
        self.PLATZART        = "  #PLATZART = \"KPS\";"
        self.ZERTIFIKATE     = "  #ZERTIFIKATE = \"Y\";"
        self.ZEITPUNKT       = "  #ZEITPUNKT = \"" + self.time_date + "\";"
        self.PRUEFKENN       = "  #PRUEFKENN = \"" + self.Result[0] + "\";"
        self.bracket         = "}"
        self.bracket2        = "  }"
        self.PRUEFUNG        = "#PRUEFUNG{"
        self.SATZ            = "  #SATZ{"
        self.EINBAUBEZ       = "    #EINBAUBEZ = \"" + self.FirmwareVersion + "\";"
        self.TESTART         = "    #TESTART = \"Pass Fail Test\";"
        self.MESSWERT        = "    #MESSWERT = \"" + self.Result.capitalize() + "\";"
        self.PRUEFUNTERSTUFE = "    #PRUEFUNTERSTUFE = \"192\";"
        self.STATUS          = "    #STATUS = \"" + self.Result[0] + "\";"
        self.EOF             = "#EOF"
        # ------------------------------------------------------
        self.nameQDMdata = self.Material + "_" + self.QRcode + "_" + self.time_date + "." + self.NameOfStation 
        self.pathofQDMdata = os.path.join(OfflineQDMdata, self.nameQDMdata)
        try:
            with open(self.pathofQDMdata, 'w') as QDMdata:
                QDMdata.write(self.qdmsender + "\n")
                QDMdata.write(self.qdmdatum + "\n")
                QDMdata.write(self.qdmlaenge + "\n")
                QDMdata.write(self.qdmprotokoll + "\n")
                QDMdata.write(self.minus + "\n")
                QDMdata.write(self.KOPF + "\n")
                QDMdata.write(self.BARCODE + "\n")
                QDMdata.write(self.MATERIAL + "\n")
                QDMdata.write(self.PRUEFCODE + "\n")
                QDMdata.write(self.PRUEFMETHODE + "\n")
                QDMdata.write(self.PRUEFFLOWCODE + "\n")
                QDMdata.write(self.PLATZNUMMER + "\n")
                QDMdata.write(self.PLATZART + "\n")
                QDMdata.write(self.ZERTIFIKATE + "\n")
                QDMdata.write(self.ZEITPUNKT + "\n")
                QDMdata.write(self.PRUEFKENN + "\n")
                QDMdata.write(self.bracket + "\n")
                QDMdata.write("\n")
                QDMdata.write(self.PRUEFUNG + "\n")
                QDMdata.write(self.SATZ + "\n")
                QDMdata.write(self.EINBAUBEZ + "\n")
                QDMdata.write(self.TESTART + "\n")
                QDMdata.write(self.MESSWERT + "\n")
                QDMdata.write(self.PRUEFUNTERSTUFE + "\n")
                QDMdata.write(self.STATUS + "\n")
                QDMdata.write(self.bracket2 + "\n")
                QDMdata.write("\n")
                QDMdata.write(self.bracket + "\n")
                QDMdata.write(self.EOF + "\n")
        except FileNotFoundError:
            pass

    def update_excel_file(self):
        TimeDay = time.ctime()

        wb = openpyxl.load_workbook(self.logviewspath) 
        sheet = wb.active 
        sheet.append((self.QRcode, self.Employee, self.Result, self.sumtime, TimeDay, self.FirmwareVersion))
        
        for cell in sheet['C']:
            if cell.value == "FAIL":
                # Set the fill color of the cell to red
                cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            elif cell.value == "PASS":
                # Set the fill color of the cell to green
                cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')    

        wb.save(self.logviewspath)
        wb.close()

    def showDialog(self, information_data):
        msgBox = QMessageBox()
        msgBox.setIcon(QMessageBox.Information)
        msgBox.setText(information_data)
        msgBox.setWindowTitle("Information / Thông tin")
        msgBox.setStandardButtons(QMessageBox.Ok)
        msgBox.exec()

    def is_file_open(self):
        try:
            os.rename(self.logviewspath, self.logviewspath)    
        except:
            os.system("taskkill /f /im excel.exe")
        return False

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    # app.setStyle('Fusion')
    # app.setStyle('windownsvista')

    window = LoginWindow()
    window.show()
    app.exec()